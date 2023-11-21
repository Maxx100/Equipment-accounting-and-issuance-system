import os
from openpyxl import Workbook, load_workbook
from datetime import datetime


class Excel:
    def __init__(self):
        if not os.path.isfile("data/log.xlsx"):
            self.wb = Workbook()
            self.wb.create_sheet("log")
            self.wb.remove_sheet(self.wb.get_sheet_by_name("Sheet"))
            self.wb.save("data/log.xlsx")
            self.sheet['A1'] = "Time"
            self.sheet['B1'] = "Date"
            self.sheet['C1'] = "Id"
            self.sheet['D1'] = "Equipment"
            self.sheet['E1'] = "Event"
            self.sheet['H1'] = 0
        else:
            self.wb = load_workbook('data/log.xlsx')
        self.read = self.wb.active
        self.sheet = self.wb['log']



    def add(self, ident, equipment, _type):
        ind = self.sheet['H1'].value
        print(self.sheet.cell(row=1, column=8).value)
        print(self.read.cell(row=1, column=8).value)
        print(ind)
        exit(0)
        self.sheet.cell(row=ind + 2, column=1).value = str(datetime.now()).split(".")[0].split()[1]
        self.sheet.cell(row=ind + 2, column=2).value = str(datetime.now()).split(".")[0].split()[0]
        self.sheet.cell(row=ind + 2, column=3).value = ident
        self.sheet.cell(row=ind + 2, column=4).value = equipment
        self.sheet.cell(row=ind + 2, column=5).value = _type
        self.sheet['H1'] = ind + 1
        self.wb.save("data/log.xlsx")
